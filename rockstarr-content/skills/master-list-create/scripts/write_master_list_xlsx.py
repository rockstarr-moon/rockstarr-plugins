#!/usr/bin/env python3
"""Write the Master List of Content workbook (.xlsx) from a rows JSON.

The skill does the messy part — reading 05_published/_publish.log and the
per-piece front-matter, joining by slug into one row per long-form piece.
This script does the deterministic part: writing a properly-formatted
Excel workbook (header styling, frozen header row, a Blog/Case study
dropdown on the Content type column). Keeping the formatting in a script
means every run produces an identical, correct workbook instead of the
agent hand-rolling openpyxl calls each time.

Usage:
    python3 write_master_list_xlsx.py --rows rows.json --out /path/master-list-of-content.xlsx
    python3 write_master_list_xlsx.py --rows rows.json --out out.xlsx --client "Acme Co"

rows.json is a JSON array of objects. Recognized keys (all optional per row;
missing values render blank):
    content_type        "Blog" | "Case study"
    linkedin_newsletter  date string, e.g. "2026-05-12" (Posted on LinkedIn newsletter)
    email_newsletter     date string (Posted in email newsletter)
    url                  URL of final blog
    title                piece title
    published_date       original publish date
    slug                 workspace slug (join key)

Column order written (canonical four first, then generated extras):
    Content type | Posted on LinkedIn newsletter | Posted in email newsletter |
    URL of final blog | Title | Published date | Slug
"""

import argparse
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "ERROR: openpyxl is required to write the .xlsx workbook.\n"
        "Install it in the Cowork Python environment:\n"
        "    python3 -m pip install openpyxl\n"
    )
    sys.exit(3)

# (header label, row-dict key). Canonical four first, then generated extras.
COLUMNS = [
    ("Content type", "content_type"),
    ("Posted on LinkedIn newsletter", "linkedin_newsletter"),
    ("Posted in email newsletter", "email_newsletter"),
    ("URL of final blog", "url"),
    ("Title", "title"),
    ("Published date", "published_date"),
    ("Slug", "slug"),
]

CONTENT_TYPES = ["Blog", "Case study"]


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--rows", required=True, help="Path to rows JSON array")
    p.add_argument("--out", required=True, help="Output .xlsx path")
    p.add_argument("--client", default="", help="Client name (informational; not required)")
    args = p.parse_args()

    with open(args.rows, encoding="utf-8") as fh:
        rows = json.load(fh)
    if not isinstance(rows, list):
        sys.stderr.write("ERROR: rows JSON must be an array of objects.\n")
        sys.exit(2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Content"

    # Header row.
    for col_idx, (label, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)

    # Data rows.
    for r_idx, row in enumerate(rows, start=2):
        for col_idx, (_label, key) in enumerate(COLUMNS, start=1):
            ws.cell(row=r_idx, column=col_idx, value=row.get(key, "") or "")

    # Freeze the header row.
    ws.freeze_panes = "A2"

    # Blog / Case study dropdown on the Content type column (column A).
    dv = DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(CONTENT_TYPES)),
        allow_blank=True,
        showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
    )
    ws.add_data_validation(dv)
    last_row = max(len(rows) + 1, 1000)  # cover headroom for manual additions
    dv.add("A2:A{}".format(last_row))

    # Reasonable column widths.
    widths = [16, 30, 28, 48, 40, 16, 28]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(args.out)
    print("Wrote {} rows to {}".format(len(rows), args.out))


if __name__ == "__main__":
    sys.exit(main() or 0)
