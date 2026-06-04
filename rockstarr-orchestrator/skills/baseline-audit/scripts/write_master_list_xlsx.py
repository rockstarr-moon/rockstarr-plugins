#!/usr/bin/env python3
"""Write the comprehensive Master List workbook (.xlsx) from a rows JSON.

This is the orchestrator-owned master list: it spans BOTH long-form
content (blogs, case studies, LinkedIn-newsletter editions, email
newsletters) AND social posts — the cross-channel "what already exists"
inventory the baseline-audit assembles. The skill does the messy part
(reading 05_published/_publish.log + per-publish records and grouping
them); this script does the deterministic part: writing a properly
formatted two-sheet workbook (header styling, frozen header rows, a
Content-type dropdown on the long-form sheet). Keeping formatting in a
script means every run produces an identical, correct workbook.

Usage:
    python3 write_master_list_xlsx.py --rows rows.json --out /path/master-list.xlsx
    python3 write_master_list_xlsx.py --rows rows.json --out out.xlsx --client "Acme Co"

rows.json is a JSON OBJECT with two arrays (each optional; missing => empty sheet):
    {
      "long_form": [ {content_type, linkedin_newsletter, email_newsletter,
                      url, title, published_date, slug}, ... ],
      "social":    [ {channel, type, published_date, permalink, topic}, ... ]
    }
All per-row keys are optional; missing values render blank. A plain JSON
array is accepted too and treated as the long_form sheet (back-compat).
"""

import argparse
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "ERROR: openpyxl is required to write the .xlsx workbook.\n"
        "Install it in the Cowork Python environment:\n"
        "    python3 -m pip install openpyxl\n"
    )
    sys.exit(3)

# (header label, row-dict key) for each sheet.
LONG_FORM_COLUMNS = [
    ("Content type", "content_type"),
    ("Posted on LinkedIn newsletter", "linkedin_newsletter"),
    ("Posted in email newsletter", "email_newsletter"),
    ("URL of final blog", "url"),
    ("Title", "title"),
    ("Published date", "published_date"),
    ("Slug", "slug"),
]
LONG_FORM_WIDTHS = [16, 30, 28, 48, 40, 16, 28]
CONTENT_TYPES = ["Blog", "Case study", "LinkedIn newsletter", "Email newsletter"]

SOCIAL_COLUMNS = [
    ("Channel", "channel"),
    ("Type", "type"),
    ("Published date", "published_date"),
    ("Permalink", "permalink"),
    ("Topic", "topic"),
]
SOCIAL_WIDTHS = [16, 18, 16, 52, 50]


def write_sheet(ws, columns, widths, rows, dropdown=None):
    for col_idx, (label, _key) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True)
    for r_idx, row in enumerate(rows, start=2):
        for col_idx, (_label, key) in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=col_idx, value=row.get(key, "") or "")
    ws.freeze_panes = "A2"
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    if dropdown:
        col_letter, choices = dropdown
        dv = DataValidation(
            type="list",
            formula1='"{}"'.format(",".join(choices)),
            allow_blank=True,
            showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
        )
        ws.add_data_validation(dv)
        last_row = max(len(rows) + 1, 1000)
        dv.add("{c}2:{c}{n}".format(c=col_letter, n=last_row))


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--rows", required=True, help="Path to rows JSON")
    p.add_argument("--out", required=True, help="Output .xlsx path")
    p.add_argument("--client", default="", help="Client name (informational)")
    args = p.parse_args()

    with open(args.rows, encoding="utf-8") as fh:
        data = json.load(fh)

    if isinstance(data, list):  # back-compat: bare array == long_form
        data = {"long_form": data, "social": []}
    if not isinstance(data, dict):
        sys.stderr.write("ERROR: rows JSON must be an object or array.\n")
        sys.exit(2)

    long_form = data.get("long_form", []) or []
    social = data.get("social", []) or []

    wb = Workbook()
    ws_lf = wb.active
    ws_lf.title = "Long-form"
    write_sheet(ws_lf, LONG_FORM_COLUMNS, LONG_FORM_WIDTHS, long_form,
                dropdown=("A", CONTENT_TYPES))

    ws_soc = wb.create_sheet("Social")
    write_sheet(ws_soc, SOCIAL_COLUMNS, SOCIAL_WIDTHS, social)

    wb.save(args.out)
    print("Wrote {} long-form + {} social rows to {}".format(
        len(long_form), len(social), args.out))


if __name__ == "__main__":
    sys.exit(main() or 0)
